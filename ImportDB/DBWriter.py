import traceback, re

from CollectionsEnum import Collections
from InitializeDB import *
from openpyxl import *


class DBWriter:
    def __init__(self, wb, db):
        self.__wb = wb
        self.__db = db
        try:
            self.__foodSheet = self.__wb[Collections.FOODS.name]
            self.__storeSheet = self.__wb[Collections.STORES.name]
        except Exception as err:
            print("Error opening sheets:", err)

    def __getID(self, *args):
        returnStr = '-'.join(args)
        return "".join(c.lower() for c in returnStr if (c.isalpha() or c in (' ', '-', "_")))

    def __getDataFromRow(self, collection, row):
        if (collection == Collections.FOODS):
            name = row[0].value
            price = row[1].value
            storeID = row[2].value
            url = row[3].value
            filters = [] if row[4].value is None else row[4].value.strip(' ').split(',')

            foodID = self.__getID(storeID, name)

            foodObj = {
                u'name': name,
                u'price': price,
                u'storeID': storeID,
                u'imageURL': url,
                u'filterKeywords': filters,
            }

            return foodID, foodObj


        elif (collection == Collections.STORES):
            store_ID = row[0].value
            store_name = row[1].value
            location = row[2].value
            open_hours = row[3].value
            close_hours = row[4].value

            storeObj = {
                u'store_name': store_name,
                u'location': location,
                u'open_hours': open_hours,
                u'close_hours': close_hours
            }

            return store_ID, storeObj

    def __readSheet(self, collection):
        try:
            sheet = None

            if (collection == Collections.FOODS):
                sheet = self.__foodSheet
            elif (collection == Collections.STORES):
                sheet = self.__storeSheet
            else:
                raise NameError("Illegal sheet name:", collection.name)

            dataSet = {}

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if (row[0].value is None):
                    break

                dataID, dataObj = self.__getDataFromRow(collection, row)
                dataSet[dataID] = dataObj

            return dataSet

        except:
            traceback.print_exc()

    def __updateFirestore(self, collection):
        try:
            print("Updating {} collection".format(collection.name))

            coll = db.collection(collection.name)

            dataSet = self.__readSheet(collection)

            for key, value in dataSet.items():
                print(key, value)
                doc = coll.document(key)
                doc.set(value)

            print("Finished updating {} collection".format(collection.name))
            print('-' * 20)
            print()

        except Exception as err:
            raise Exception(err)

    def writeToFood(self):
        try:
            self.__updateFirestore(Collections.FOODS)
        except:
            traceback.print_exc()

    def writeToStore(self):
        try:
            dataSet = self.__updateFirestore(Collections.STORES)
        except:
            traceback.print_exc()


if __name__ == '__main__':
    try:
        db = InitializeDB()
        wb = load_workbook(filename="./Utown Food Menu Dataset.xlsx")
        writer = DBWriter(wb, db)
        writer.writeToFood()
        writer.writeToStore()

    except:
        traceback.print_exc()
